import json
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK_PATH = Path(r"D:\DATA_MASTER_ Students_Staff\UP_UNIT_MAPPING_REVIEW_2026-05-22.xlsx")
OUTPUT_PATH = Path(r"D:\DATA_MASTER_ Students_Staff\approved_unit_mapping_review.json")


def read_sheet_rows(workbook_path: Path, sheet_name: str):
    workbook = load_workbook(workbook_path, data_only=True)
    sheet = workbook[sheet_name]
    headers = [cell.value for cell in sheet[1]]
    rows = []
    for row_idx in range(2, sheet.max_row + 1):
        item = {
            headers[col_idx - 1]: sheet.cell(row_idx, col_idx).value
            for col_idx in range(1, len(headers) + 1)
        }
        if item.get("match_status") == "REVIEW":
            rows.append(item)
    return rows


def main():
    rows = read_sheet_rows(WORKBOOK_PATH, "Needs Review")
    actions = {}
    for row in rows:
        action = row.get("proposed_action") or ""
        actions[action] = actions.get(action, 0) + 1

    missing_action = [
        index + 2 for index, row in enumerate(rows) if not row.get("proposed_action")
    ]
    missing_approved = [
        index + 2
        for index, row in enumerate(rows)
        if row.get("proposed_action") in ("use_existing", "map_to_parent")
        and not row.get("approved_unit_id")
    ]
    missing_new_unit_id = [
        index + 2
        for index, row in enumerate(rows)
        if row.get("proposed_action") in ("add_active_unit", "add_inactive_unit")
        and not (row.get("approved_unit_id") or row.get("proposed_unit_id"))
    ]

    output = {
        "source_workbook": str(WORKBOOK_PATH),
        "privacy_note": "Aggregate unit mapping only. No person-level rows, names, or IDs are included.",
        "row_count": len(rows),
        "actions": actions,
        "missing_action_rows": missing_action,
        "missing_approved_rows": missing_approved,
        "missing_new_unit_id_rows": missing_new_unit_id,
        "review_rows": rows,
    }

    OUTPUT_PATH.write_text(
        json.dumps(output, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(str(OUTPUT_PATH))
    print(json.dumps({k: output[k] for k in ("row_count", "actions", "missing_action_rows", "missing_approved_rows", "missing_new_unit_id_rows")}, ensure_ascii=False))


if __name__ == "__main__":
    main()
